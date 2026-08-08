#!/usr/bin/env python3
"""
Unit tests for hikvision_fetch.py — the Hikvision ISAPI attendance fetch script.

These tests never touch the network: HTTP behavior is exercised through mocked
session responses. Run from the repo root:

    .venv/bin/python -m pytest UkuuHr.Tests/test_hikvision_fetch.py -v

or just `pytest` (the file matches pytest's default `test_*.py` discovery).
"""

import datetime as dt
import json
import os
import re
import subprocess
import sys

import pytest
import requests

# Make the repo-root script importable regardless of the CWD.
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

import hikvision_fetch as hf  # noqa: E402


# ─────────────────────────────────────────────────────────────────────────────
# Time parsing
# ─────────────────────────────────────────────────────────────────────────────

class TestParseHikTime:
    def test_iso_with_offset(self):
        parsed = hf._parse_hik_time("2024-07-25T08:30:00+08:00")
        assert parsed.isoformat() == "2024-07-25T00:30:00+00:00"

    def test_iso_utc_z(self):
        parsed = hf._parse_hik_time("2024-07-25T08:30:00Z")
        assert parsed.isoformat() == "2024-07-25T08:30:00+00:00"

    def test_space_separated(self):
        parsed = hf._parse_hik_time("2024-07-25 08:30:00")
        assert parsed.isoformat() == "2024-07-25T08:30:00+00:00"

    def test_compact_format(self):
        parsed = hf._parse_hik_time("20240725T083000+0800")
        assert parsed.isoformat() == "2024-07-25T00:30:00+00:00"

    def test_fractional_seconds(self):
        parsed = hf._parse_hik_time("2024-07-25T08:30:00.123+08:00")
        assert parsed.isoformat() == "2024-07-25T00:30:00.123000+00:00"

    def test_empty_raises(self):
        with pytest.raises(ValueError):
            hf._parse_hik_time("")

    def test_garbage_raises(self):
        with pytest.raises(ValueError):
            hf._parse_hik_time("not-a-time")


# ─────────────────────────────────────────────────────────────────────────────
# XML helpers (namespace-insensitive)
# ─────────────────────────────────────────────────────────────────────────────

class TestXmlHelpers:
    @staticmethod
    def _ns_root():
        return hf.ET.fromstring(
            '<?xml version="1.0"?>'
            '<DeviceInfo xmlns="http://www.hikvision.com/ver20/XMLSchema">'
            "<deviceName>DS-K1T671</deviceName><model>DS-K1T671M</model>"
            "</DeviceInfo>"
        )

    def test_local_name_strips_namespace(self):
        assert hf._local_name("{http://www.hikvision.com/ver20/XMLSchema}deviceName") == "deviceName"
        assert hf._local_name("plain") == "plain"

    def test_find_local_ignores_namespace(self):
        root = self._ns_root()
        el = hf._find_local(root, "deviceName")
        assert el is not None and el.text == "DS-K1T671"

    def test_find_local_missing(self):
        assert hf._find_local(self._ns_root(), "nope") is None

    def test_find_all_local(self):
        root = hf.ET.fromstring(
            '<?xml version="1.0"?><Response xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><a>1</a></LogItem><LogItem><a>2</a></LogItem></Response>"
        )
        items = hf._find_all_local(root, "LogItem")
        assert len(items) == 2


# ─────────────────────────────────────────────────────────────────────────────
# AcsEvent JSON parsing
# ─────────────────────────────────────────────────────────────────────────────

class TestParseAcsEventJson:
    def test_info_list_with_checkout(self):
        sample = json.dumps({
            "AcsEvent": {
                "searchID": "s1",
                "numOfMatches": 1,
                "totalMatches": 1,
                "InfoList": [{
                    "employeeNo": "EMP001",
                    "time": "2024-07-25T08:30:00+08:00",
                    "major": 1,
                    "minor": 76,
                    "verifyMode": 15,
                    "inAndOutMode": 1,
                }],
            }
        })
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert parsed["num_of_matches"] == 1
        assert parsed["total_matches"] == 1
        assert len(parsed["events"]) == 1
        evt = parsed["events"][0]
        assert evt["employee_no"] == "EMP001"
        assert evt["event_type"] == "CheckOut"
        assert evt["major"] == 1 and evt["minor"] == 76

    def test_event_list_checkin(self):
        sample = json.dumps({
            "AcsEvent": {
                "numOfMatches": 2,
                "totalMatches": 2,
                "EventList": [
                    {"employeeNo": "E1", "time": "2024-07-25T09:00:00+08:00", "major": 1, "minor": 75},
                    {"employeeNo": "E2", "time": "2024-07-25 10:00:00", "major": 1, "minor": 76},
                ],
            }
        })
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert len(parsed["events"]) == 2
        assert parsed["events"][0]["event_type"] == "CheckIn"
        assert parsed["events"][1]["event_type"] == "CheckOut"

    def test_top_level_info_list(self):
        sample = json.dumps({"InfoList": [{"employeeNo": "E9", "time": "2024-07-25T09:00:00Z",
                                           "major": 1, "minor": 75}]})
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert len(parsed["events"]) == 1
        assert parsed["events"][0]["employee_no"] == "E9"

    def test_access_events_classified_other(self):
        """Door/card/access events (major != 1) must NOT be reported as
        CheckIn/CheckOut — they are access events, not attendance punches.
        This mirrors what a real DS-K1T321MFWX emits for card reads."""
        sample = json.dumps({
            "AcsEvent": {"InfoList": [
                {"employeeNo": "A", "time": "2024-07-25T09:00:00Z",
                 "major": 5, "minor": 9, "cardNo": "1234"},
                {"employeeNo": "B", "time": "2024-07-25T09:05:00Z",
                 "major": 5, "minor": 21, "doorNo": 1},
                {"employeeNo": "C", "time": "2024-07-25T09:10:00Z",
                 "major": 3, "minor": 1028},
            ]}
        })
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert len(parsed["events"]) == 3
        assert all(e["event_type"] == "Other" for e in parsed["events"])

    def test_classify_event_helper(self):
        assert hf._classify_event(1, 75) == "CheckIn"
        assert hf._classify_event(1, 76) == "CheckOut"
        assert hf._classify_event(5, 9) == "Other"
        assert hf._classify_event(5, 21) == "Other"
        assert hf._classify_event(3, 1028) == "Other"
        assert hf._classify_event(0, 0) == "Other"

    def test_no_match(self):
        parsed = hf.HikvisionClient._parse_acs_event_json(
            '{"AcsEvent": {"searchID": "s30", "totalMatches": 0, '
            '"responseStatusStrg": "NO MATCH", "numOfMatches": 0}}')
        assert parsed["events"] == []
        assert parsed["total_matches"] == 0

    def test_invalid_json_returns_empty(self):
        parsed = hf.HikvisionClient._parse_acs_event_json("not json {{")
        assert parsed["events"] == []
        assert parsed["num_of_matches"] == 0

    def test_skips_unparseable_time(self):
        sample = json.dumps({
            "AcsEvent": {"InfoList": [
                {"employeeNo": "A", "time": "2024-07-25T09:00:00Z", "major": 1, "minor": 75},
                {"employeeNo": "B", "time": "garbage", "major": 1, "minor": 76},
            ]}
        })
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert len(parsed["events"]) == 1
        assert parsed["events"][0]["employee_no"] == "A"
        assert parsed["skipped"] == 1

    def test_counts_non_dict_items_as_skipped(self):
        sample = json.dumps({
            "AcsEvent": {"InfoList": [
                {"employeeNo": "A", "time": "2024-07-25T09:00:00Z", "major": 1, "minor": 75},
                "not-a-dict",
                None,
            ]}
        })
        parsed = hf.HikvisionClient._parse_acs_event_json(sample)
        assert len(parsed["events"]) == 1
        assert parsed["skipped"] == 2

    def test_invalid_json_reports_zero_skipped(self):
        parsed = hf.HikvisionClient._parse_acs_event_json("not json {{")
        assert parsed["skipped"] == 0


# ─────────────────────────────────────────────────────────────────────────────
# AuditLog XML parsing
# ─────────────────────────────────────────────────────────────────────────────

class TestParseAuditLogXml:
    def test_namespaced_log_items(self):
        xml_text = (
            '<?xml version="1.0"?>'
            '<AuditLogSearchResponse xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><employeeNo>E2</employeeNo><time>2024-07-25T09:00:00+08:00</time>"
            "<major>1</major><minor>75</minor></LogItem>"
            "<LogItem><employeeNo>E3</employeeNo><time>2024-07-25T10:00:00+08:00</time>"
            "<major>1</major><minor>76</minor></LogItem>"
            "</AuditLogSearchResponse>"
        )
        events = hf.HikvisionClient._parse_audit_log_xml(xml_text)
        assert len(events) == 2
        assert events[0]["employee_no"] == "E2"
        assert events[0]["event_type"] == "CheckIn"
        assert events[1]["event_type"] == "CheckOut"

    def test_invalid_xml_returns_empty(self):
        assert hf.HikvisionClient._parse_audit_log_xml("<<<not xml>>>") == []

    def test_non_attendance_log_items_classified_other(self):
        """AuditLog entries with unknown major/minor codes are Other, not CheckIn."""
        xml_text = (
            '<?xml version="1.0"?>'
            '<AuditLogSearchResponse xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><employeeNo>E4</employeeNo><time>2024-07-25T09:00:00+08:00</time>"
            "<major>5</major><minor>9</minor></LogItem>"
            "</AuditLogSearchResponse>"
        )
        events = hf.HikvisionClient._parse_audit_log_xml(xml_text)
        assert len(events) == 1
        assert events[0]["event_type"] == "Other"


# ─────────────────────────────────────────────────────────────────────────────
# Timezone offset formatting
# ─────────────────────────────────────────────────────────────────────────────

class TestDeviceTimezone:
    def test_fmt_hik_time_converts_to_device_local(self):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x", tz_offset="+08:00")
        utc = dt.datetime(2024, 7, 25, 0, 30, tzinfo=dt.timezone.utc)
        assert client._fmt_hik_time(utc) == "2024-07-25T08:30:00+08:00"

    def test_fmt_hik_time_negative_offset(self):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x", tz_offset="-05:00")
        utc = dt.datetime(2024, 7, 25, 12, 0, tzinfo=dt.timezone.utc)
        assert client._fmt_hik_time(utc) == "2024-07-25T07:00:00-05:00"

    def test_fmt_hik_time_naive_means_utc(self):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x", tz_offset="+00:00")
        naive = dt.datetime(2024, 7, 25, 8, 30)
        assert client._fmt_hik_time(naive) == "2024-07-25T08:30:00+00:00"

    def test_get_device_offset_autodetect(self):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x")

        class FakeResp:
            text = '<?xml version="1.0"?><Time xmlns="http://www.isapi.org/ver20/XMLSchema">' \
                   "<localTime>1970-01-01T08:35:16+02:00</localTime></Time>"

        client.get = lambda path: FakeResp()  # noqa: E731
        assert client._get_device_offset() == "+02:00"
        # cached on second call
        client.get = lambda path: (_ for _ in ()).throw(AssertionError("should be cached"))  # noqa: E731
        assert client._get_device_offset() == "+02:00"

    def test_get_device_offset_fallback(self, capsys):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x")
        client.get = lambda path: (_ for _ in ()).throw(OSError("boom"))  # noqa: E731
        assert client._get_device_offset() == "+08:00"
        assert "WARNING" in capsys.readouterr().err


# ─────────────────────────────────────────────────────────────────────────────
# AcsEvent search + pagination (mocked HTTP)
# ─────────────────────────────────────────────────────────────────────────────

class FakeResponse:
    def __init__(self, text, status_code=200):
        self.text = text
        self.status_code = status_code

    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError(f"HTTP {self.status_code}")


class FakeClient(hf.HikvisionClient):
    """A HikvisionClient whose session is replaced by a scripted responder."""

    def __init__(self, responses_by_path):
        super().__init__("1.2.3.4", "admin", "x", tz_offset="+08:00")
        self._responses = responses_by_path
        self.request_log = []

    def get(self, path):
        self.request_log.append(("GET", path))
        for prefix, resp in self._responses.items():
            if path.startswith(prefix):
                return resp
        raise AssertionError(f"unexpected GET {path}")

    def post(self, path, body, content_type="application/xml"):
        self.request_log.append(("POST", path, body))
        parsed = json.loads(body)
        cond = parsed.get("AcsEventCond")
        assert cond is not None, "JSON body must be wrapped in AcsEventCond"
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}[+-]\d{2}:\d{2}",
                            cond["startTime"]), "startTime must carry a numeric UTC offset"
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}[+-]\d{2}:\d{2}",
                            cond["endTime"]), "endTime must carry a numeric UTC offset"
        # Responses may be keyed by searchResultPosition (int) for multi-page
        # scripts, or by path prefix (str) for single-page ones.
        pos = cond.get("searchResultPosition")
        if pos in self._responses:
            return self._responses[pos]
        for prefix, resp in self._responses.items():
            if isinstance(prefix, str) and path.startswith(prefix):
                return resp
        raise AssertionError(f"unexpected POST {path} (position={pos})")


class TestAcsEventSearch:
    def test_single_page(self):
        page = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 1, "totalMatches": 1,
                                        "InfoList": [{"employeeNo": "E1",
                                                      "time": "2024-07-25T09:00:00+08:00",
                                                      "major": 1, "minor": 75}]}})
        client = FakeClient({"/ISAPI/AccessControl/AcsEvent": FakeResponse(page)})
        since = dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc)
        events = client._acs_event_search(since, max_results=1000)
        assert len(events) == 1
        assert events[0]["employee_no"] == "E1"
        # Only one request: single page, no continuation.
        assert len(client.request_log) == 1

    def test_pagination_multiple_pages(self):
        page1 = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 2, "totalMatches": 4,
                                         "InfoList": [{"employeeNo": "A",
                                                       "time": "2024-07-25T09:00:00+08:00",
                                                       "major": 1, "minor": 75},
                                                      {"employeeNo": "B",
                                                       "time": "2024-07-25T09:05:00+08:00",
                                                       "major": 1, "minor": 75}]}})
        page2 = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 2, "totalMatches": 4,
                                         "InfoList": [{"employeeNo": "C",
                                                       "time": "2024-07-25T09:10:00+08:00",
                                                       "major": 1, "minor": 75},
                                                      {"employeeNo": "D",
                                                       "time": "2024-07-25T09:15:00+08:00",
                                                       "major": 1, "minor": 75}]}})
        client = FakeClient({0: FakeResponse(page1), 2: FakeResponse(page2)})
        since = dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc)
        events = client._acs_event_search(since, max_results=1000)
        assert [e["employee_no"] for e in events] == ["A", "B", "C", "D"]
        assert len(client.request_log) == 2
        # Continuation pages are POSTs carrying the same searchID with an
        # incremented searchResultPosition (GET continuations 404 on the
        # DS-K1T321MFWX firmware this mirrors).
        assert client.request_log[1][0] == "POST"
        cond0 = json.loads(client.request_log[0][2])["AcsEventCond"]
        cond1 = json.loads(client.request_log[1][2])["AcsEventCond"]
        assert cond0["searchID"] == cond1["searchID"]
        assert cond0["searchResultPosition"] == 0
        assert cond1["searchResultPosition"] == 2

    def test_zero_matches_terminates(self):
        no_match = json.dumps({"AcsEvent": {"searchID": "s0", "totalMatches": 0,
                                            "responseStatusStrg": "NO MATCH", "numOfMatches": 0}})
        client = FakeClient({"/ISAPI/AccessControl/AcsEvent": FakeResponse(no_match)})
        events = client._acs_event_search(None, max_results=1000)
        assert events == []
        assert len(client.request_log) == 1

    def test_attendance_only_asks_for_major_1_and_filters(self):
        """--attendance-only must (a) send major=1 in the AcsEventCond body and
        (b) drop any non-attendance events as a client-side safety net."""
        page = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 3, "totalMatches": 3,
                                        "InfoList": [
                                            {"employeeNo": "A",
                                             "time": "2024-07-25T09:00:00+08:00",
                                             "major": 1, "minor": 75},
                                            {"employeeNo": "B",
                                             "time": "2024-07-25T09:05:00+08:00",
                                             "major": 5, "minor": 9},  # card read - filtered
                                            {"employeeNo": "C",
                                             "time": "2024-07-25T09:10:00+08:00",
                                             "major": 1, "minor": 76},
                                        ]}})
        client = FakeClient({0: FakeResponse(page)})
        since = dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc)
        events = client._acs_event_search(since, max_results=100, attendance_only=True)
        assert [e["employee_no"] for e in events] == ["A", "C"]
        assert client.filtered_events == 1
        cond = json.loads(client.request_log[0][2])["AcsEventCond"]
        assert cond["major"] == 1

    def test_default_search_uses_major_0_and_keeps_all(self):
        """Without --attendance-only, major=0 fetches everything (unchanged)."""
        page = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 2, "totalMatches": 2,
                                        "InfoList": [
                                            {"employeeNo": "A",
                                             "time": "2024-07-25T09:00:00+08:00",
                                             "major": 5, "minor": 9},
                                            {"employeeNo": "B",
                                             "time": "2024-07-25T09:05:00+08:00",
                                             "major": 1, "minor": 75},
                                        ]}})
        client = FakeClient({0: FakeResponse(page)})
        events = client._acs_event_search(None, max_results=100, attendance_only=False)
        assert [e["employee_no"] for e in events] == ["A", "B"]
        assert client.filtered_events == 0
        cond = json.loads(client.request_log[0][2])["AcsEventCond"]
        assert cond["major"] == 0

    def test_attendance_only_audit_log_fallback(self):
        """The AuditLog fallback also applies the major=1 filter."""
        audit_xml = (
            '<?xml version="1.0"?>'
            '<AuditLogSearchResponse xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><employeeNo>F1</employeeNo><time>2024-07-25T09:00:00+08:00</time>"
            "<major>1</major><minor>75</minor></LogItem>"
            "<LogItem><employeeNo>F2</employeeNo><time>2024-07-25T09:05:00+08:00</time>"
            "<major>5</major><minor>9</minor></LogItem>"
            "</AuditLogSearchResponse>"
        )
        client = FakeClient({"/ISAPI/AccessControl/AuditLog/search": FakeResponse(audit_xml)})
        since = dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc)
        events = client._audit_log_search(since, attendance_only=True)
        assert [e["employee_no"] for e in events] == ["F1"]
        assert client.filtered_events == 1


# ─────────────────────────────────────────────────────────────────────────────
# Request behavior: fail-fast on connection errors, JSON body wiring
# ─────────────────────────────────────────────────────────────────────────────

class TestRequestBehavior:
    def test_connection_error_fails_fast(self):
        class Boom:
            def request(self, *a, **k):
                raise requests.exceptions.ConnectionError("refused")

        client = hf.HikvisionClient("1.2.3.4", "admin", "x")
        client.session = Boom()
        with pytest.raises(ConnectionError, match="Cannot connect"):
            client.get("/ISAPI/System/deviceInfo")

    def test_timeout_tuple_configured(self):
        client = hf.HikvisionClient("1.2.3.4", "admin", "x", timeout=25)
        assert client.timeout == (hf.CONNECT_TIMEOUT, 25)

    def test_digest_auth_retry_on_401(self):
        """The core fix for the DS-K1T321MFWX: Basic gets 401, the client must
        retry the same request with digest auth and return the 200 response."""
        basic_401 = FakeResponse('{"statusCode":4}', status_code=401)
        digest_ok = FakeResponse('{"AcsEvent":{"numOfMatches":0,"totalMatches":0}}')

        class ScriptedSession:
            def __init__(self):
                self.calls = []

            def request(self, method, url, auth=None, **kwargs):
                self.calls.append((method, url, type(auth).__name__))
                if isinstance(auth, tuple):  # Basic auth tuple -> reject
                    return basic_401
                return digest_ok  # digest -> accept

        session = ScriptedSession()
        client = hf.HikvisionClient("1.2.3.4", "admin", "x", tz_offset="+08:00")
        client.session = session

        resp = client.get("/ISAPI/System/deviceInfo")
        assert resp.status_code == 200
        assert session.calls[0][2] == "tuple"  # first attempt Basic
        assert session.calls[1][2] == "HTTPDigestAuth"  # retry with digest

    def test_audit_log_fallback_counts_skipped(self):
        """The AuditLog fallback also accumulates the skipped-event counter when
        raw LogItems fail to parse (e.g. unparseable time)."""
        audit_xml = (
            '<?xml version="1.0"?>'
            '<AuditLogSearchResponse xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><employeeNo>F1</employeeNo><time>2024-07-25T09:00:00+08:00</time>"
            "<major>1</major><minor>75</minor></LogItem>"
            "<LogItem><employeeNo>F2</employeeNo><time>garbage</time>"
            "<major>1</major><minor>76</minor></LogItem>"
            "</AuditLogSearchResponse>"
        )
        client = FakeClient({"/ISAPI/AccessControl/AuditLog/search": FakeResponse(audit_xml)})
        events = client._audit_log_search(dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc))
        assert len(events) == 1
        assert client.skipped_events == 1

    def test_safety_cap_triggers_warning_and_flag(self, capsys):
        """A misbehaving device that never reports totalMatches hits the safety
        cap: the flag is set and a stderr warning explains the truncation."""
        # totalMatches is huge and never decreases, so pagination never converges
        # naturally -> the safety cap (max_results * 20) must fire.
        page = json.dumps({"AcsEvent": {"searchID": "s1", "numOfMatches": 100,
                                        "totalMatches": 999999,
                                        "InfoList": [{"employeeNo": "A",
                                                      "time": "2024-07-25T09:00:00+08:00",
                                                      "major": 1, "minor": 75}]}})
        client = FakeClient({"/ISAPI/AccessControl/AcsEvent": FakeResponse(page)})
        events = client._acs_event_search(None, max_results=100)
        assert client.truncated_cap_hit is True
        assert "pagination safety cap" in capsys.readouterr().err
        # Position advances by numOfMatches (100) per page; cap = max_results * 20
        # = 2000 -> exactly 20 iterations, one event per page.
        assert len(events) == 20

    def test_audit_log_fallback(self):
        """fetch_attendance_events falls back to the AuditLog endpoint when the
        AcsEvent search fails, and reports which source succeeded."""
        audit_xml = (
            '<?xml version="1.0"?>'
            '<AuditLogSearchResponse xmlns="http://www.hikvision.com/ver10/XMLSchema">'
            "<LogItem><employeeNo>F1</employeeNo><time>2024-07-25T09:00:00+08:00</time>"
            "<major>1</major><minor>75</minor></LogItem></AuditLogSearchResponse>"
        )

        class FailingAcsClient(FakeClient):
            def post(self, path, body, content_type="application/xml"):
                self.request_log.append(("POST", path))
                raise requests.exceptions.HTTPError("401 Invalid Operation")

        client = FailingAcsClient({"/ISAPI/AccessControl/AuditLog/search": FakeResponse(audit_xml)})
        since = dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc)
        events, source = client.fetch_attendance_events(since=since, max_results=100)
        assert source == "AuditLog"
        assert len(events) == 1
        assert events[0]["employee_no"] == "F1"

    def test_both_endpoints_fail_raises_with_causes(self):
        class FailingClient(FakeClient):
            def post(self, path, body, content_type="application/xml"):
                raise requests.exceptions.HTTPError("AcsEvent boom")

        # The AuditLog fallback GET hits the empty responses dict, so FakeClient.get
        # raises AssertionError, which fetch_attendance_events treats as the
        # second failure and chains into the raised ConnectionError.
        client = FailingClient({})
        with pytest.raises(ConnectionError, match="AcsEvent boom"):
            client.fetch_attendance_events(
                since=dt.datetime(2024, 7, 24, tzinfo=dt.timezone.utc), max_results=100)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

class TestExportToExcel:
    def test_export_roundtrip(self, tmp_path):
        events = [{
            "employee_no": "E1",
            "time": dt.datetime(2024, 7, 25, 9, 0, tzinfo=dt.timezone.utc),
            "event_type": "CheckIn",
            "major": 1,
            "minor": 75,
            "verify_mode": "15",
            "in_out_mode": "1",
            "raw": "<raw/>",
        }]
        out = tmp_path / "out.xlsx"
        path = hf.export_to_excel(events, str(out))
        assert os.path.exists(path)

        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.title == "Attendance Events"
        assert ws.cell(row=1, column=2).value == "Employee No"
        assert ws.cell(row=2, column=2).value == "E1"
        assert ws.cell(row=2, column=4).value == "CheckIn"


# ─────────────────────────────────────────────────────────────────────────────
# CLI validation
# ─────────────────────────────────────────────────────────────────────────────

class TestCliValidation:
    def test_bad_tz_offset_rejected(self):
        result = _run_cli(["--host", "1.2.3.4", "--tz-offset", "8:00"])
        assert result.returncode != 0
        assert "--tz-offset must look like" in result.stderr

    def test_zero_timeout_rejected(self):
        result = _run_cli(["--host", "1.2.3.4", "--timeout", "0"])
        assert result.returncode != 0
        assert "--timeout must be a positive" in result.stderr

    def test_negative_retries_rejected(self):
        result = _run_cli(["--host", "1.2.3.4", "--retries", "-1"])
        assert result.returncode != 0
        assert "--retries cannot be negative" in result.stderr

    def test_tz_offset_regex_matches_valid_offsets(self):
        assert hf.re.fullmatch(r"[+-]\d{2}:\d{2}", "+08:00")
        assert hf.re.fullmatch(r"[+-]\d{2}:\d{2}", "-05:00")
        assert not hf.re.fullmatch(r"[+-]\d{2}:\d{2}", "8:00")

    def test_https_defaults_to_port_443(self):
        result = _run_cli(["--host", "127.0.0.1", "--https", "--retries", "0"])
        assert result.returncode != 0
        # The connection attempt must target https://127.0.0.1:443.
        assert "127.0.0.1:443" in result.stderr

    def test_plain_http_defaults_to_port_80(self):
        result = _run_cli(["--host", "127.0.0.1", "--retries", "0"])
        assert result.returncode != 0
        assert "127.0.0.1:80" in result.stderr


def _run_cli(argv):
    """Run hikvision_fetch.py with the given argv in a subprocess."""
    return subprocess.run(
        [sys.executable, os.path.join(REPO_ROOT, "hikvision_fetch.py")] + argv,
        capture_output=True, text=True, timeout=60)
