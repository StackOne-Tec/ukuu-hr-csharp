#!/usr/bin/env python3
"""
hikvision_fetch.py — Fetch attendance events from a Hikvision ISAPI time & attendance
terminal, mirroring the logic in UkuuHr.Web/Services/Hikvision/HikvisionIsapiClient.cs.

Note: event classification intentionally diverges from the C# client. The C#
ParseAcsEventJson treats any non-(1,76) event as CheckIn, which mislabels access
events (door opens, card reads, alarms). Here only (1,75)=CheckIn and (1,76)=CheckOut;
everything else is "Other".

Endpoints used:
  GET  /ISAPI/System/deviceInfo                          -> device info (XML)
  POST /ISAPI/AccessControl/AcsEvent?format=json         -> attendance events (JSON)
  GET  /ISAPI/AccessControl/AuditLog/search              -> fallback (XML)

Usage:
  python hikvision_fetch.py --host 192.168.1.137 --user admin --password <device-password> --device-info
  python hikvision_fetch.py --host 192.168.1.137 --user admin --password <device-password>   # last 7 days
  python hikvision_fetch.py --host 192.168.1.137 --user admin --password <device-password> --days 30 --output events.xlsx
  python hikvision_fetch.py --host 192.168.1.137 --user admin --password <device-password> --attendance-only   # punches only

Requires: pip install requests openpyxl
"""

import argparse
import datetime as dt
import json
import re
import sys
import time
import uuid
import xml.etree.ElementTree as ET

import requests

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# ISAPI client (mirrors HikvisionIsapiClient.cs)
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_TIMEOUT = 30  # per-request read timeout (seconds)
CONNECT_TIMEOUT = 5   # connect timeout (seconds) - fail fast on unreachable hosts
MAX_RETRIES = 3
RETRY_DELAY_MS = 1000


def _local_name(tag):
    """Strip an XML namespace prefix: '{http://...}deviceName' -> 'deviceName'."""
    return tag.rsplit("}", 1)[-1]


def _find_local(root, name):
    """Find the first descendant element by local name, ignoring namespaces."""
    for el in root.iter():
        if _local_name(el.tag) == name:
            return el
    return None


def _find_all_local(root, name):
    """Find all descendant elements by local name, ignoring namespaces."""
    return [el for el in root.iter() if _local_name(el.tag) == name]


class HikvisionClient:
    def __init__(self, host, username, password, port=80, use_https=False,
                 timeout=DEFAULT_TIMEOUT, max_retries=MAX_RETRIES, tz_offset=None):
        self.base_url = f"{'https' if use_https else 'http'}://{host}:{port}"
        self.username = username
        self.password = password
        # (connect, read) tuple: unreachable hosts fail in seconds, slow device
        # responses still get a generous read window.
        self.timeout = (CONNECT_TIMEOUT, timeout)
        self.max_retries = max_retries
        self.session = requests.Session()
        self.session.headers["Accept"] = "application/xml, application/json"
        self.session.verify = False  # Hikvision devices ship with self-signed certs
        self._basic_auth = (username, password)
        # Device local-timezone offset (e.g. "+08:00"). Auto-detected from
        # /ISAPI/System/time on first use unless explicitly provided.
        self._tz_offset = tz_offset
        self._tz_timedelta = None
        # Diagnostic counters, surfaced by the CLI after a fetch.
        self.skipped_events = 0
        self.truncated_cap_hit = False
        self.filtered_events = 0

    # ----- core request helpers with retry -----

    def _request(self, method, path, **kwargs):
        kwargs.setdefault("timeout", self.timeout)
        url = f"{self.base_url}{path}"
        last_exc = None
        digest_auth = None
        from requests.auth import HTTPDigestAuth
        digest_auth = HTTPDigestAuth(self.username, self.password)
        for attempt in range(self.max_retries + 1):
            try:
                resp = self.session.request(method, url, auth=self._basic_auth, **kwargs)
                if resp.status_code == 401:
                    # Hikvision sometimes requires digest auth; use it once.
                    resp = self.session.request(method, url, auth=digest_auth, **kwargs)
                resp.raise_for_status()
                return resp
            except requests.exceptions.ConnectionError:
                # Unreachable host / refused - retrying won't help; surface fast.
                raise ConnectionError(
                    f"Cannot connect to {self.base_url}{path}") from None
            except Exception as e:  # noqa: BLE001
                last_exc = e
                if attempt < self.max_retries:
                    time.sleep(RETRY_DELAY_MS / 1000.0 * (attempt + 1))
        raise ConnectionError(f"ISAPI request failed for {self.base_url}{path}: {last_exc}")

    def get(self, path):
        return self._request("GET", path)

    def post(self, path, body, content_type="application/xml"):
        return self._request("POST", path, data=body.encode("utf-8"),
                             headers={"Content-Type": content_type})

    # ----- device info -----

    def fetch_device_info(self):
        """GET /ISAPI/System/deviceInfo -> dict (mirrors ParseDeviceInfoXml)."""
        resp = self.get("/ISAPI/System/deviceInfo")
        root = ET.fromstring(resp.text)
        def text(name):
            el = _find_local(root, name)
            return el.text.strip() if el is not None and el.text else ""
        return {
            "deviceName": text("deviceName"),
            "deviceID": text("deviceID"),
            "model": text("model"),
            "serialNumber": text("serialNumber"),
            "macAddress": text("macAddress"),
            "firmwareVersion": text("firmwareVersion"),
            "hardwareVersion": text("hardwareVersion"),
            "deviceType": text("deviceType"),
            "systemTime": text("systemTime"),
        }

    # ----- device local timezone offset -----

    def _get_device_offset(self):
        """Return the device's local UTC offset as a '+HH:MM' string.

        Auto-detects from /ISAPI/System/time (<localTime> ISO-8601 with offset),
        falling back to +08:00 (typical for Hikvision access controllers).
        """
        if self._tz_offset is not None:
            return self._tz_offset
        offset = None
        try:
            resp = self.get("/ISAPI/System/time")
            root = ET.fromstring(resp.text)
            el = _find_local(root, "localTime")
            if el is not None and el.text:
                m = re.search(r"([+-]\d{2}:\d{2})$", el.text.strip())
                if m:
                    offset = m.group(1)
        except Exception:
            offset = None
        if offset is None:
            offset = "+08:00"
            print("WARNING: could not detect device timezone offset, defaulting to +08:00",
                  file=sys.stderr)
        self._tz_offset = offset
        return offset

    def _device_tz(self):
        """Return a timezone object for the device's local offset."""
        if self._tz_timedelta is None:
            off = self._get_device_offset()
            sign = 1 if off.startswith("+") else -1
            hh, mm = off[1:].split(":")
            self._tz_timedelta = sign * dt.timedelta(hours=int(hh), minutes=int(mm))
        return dt.timezone(self._tz_timedelta)

    def _fmt_hik_time(self, value):
        """Format an aware UTC datetime as 'YYYY-MM-DDTHH:MM:SS+08:00' using the
        device's local offset (the format this device family requires)."""
        if value.tzinfo is None:
            value = value.replace(tzinfo=dt.timezone.utc)
        local = value.astimezone(self._device_tz())
        return local.strftime("%Y-%m-%dT%H:%M:%S") + self._get_device_offset()

    # ----- attendance events -----

    def fetch_attendance_events(self, since=None, max_results=1000, attendance_only=False):
        """Try AcsEvent JSON search (with pagination), fall back to AuditLog XML search."""
        # Reset per-attempt counters so a partial AcsEvent failure can't leak
        # filtered/skipped counts into the AuditLog fallback.
        self.filtered_events = 0
        self.skipped_events = 0
        try:
            events = self._acs_event_search(since, max_results, attendance_only)
            return events, "AcsEvent"
        except Exception as acs_err:
            self.filtered_events = 0
            self.skipped_events = 0
            try:
                events = self._audit_log_search(since, attendance_only)
                return events, "AuditLog"
            except Exception as audit_err:
                raise ConnectionError(
                    f"Both AcsEvent and AuditLog endpoints failed (AcsEvent: {acs_err}; "
                    f"AuditLog: {audit_err}).") from audit_err

    def _acs_event_search(self, since, max_results, attendance_only=False):
        """POST the AcsEvent search (JSON body), then follow searchID pages.

        The DS-K1T3xx / K1T321M family requires: digest auth, a JSON body
        wrapped in "AcsEventCond", and start/end times carrying a numeric UTC
        offset ("+08:00") instead of "Z".
        """
        start = since if since else (dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=7))
        end = dt.datetime.now(dt.timezone.utc)
        search_id = f"AcsEventSearch_{uuid.uuid4().hex}"

        events = []
        position = 0
        total_matches = None

        while True:
            # NOTE: every page — including continuations — must be POSTed with the
            # full AcsEventCond body. The DS-K1T321MFWX (V3.9.2) returns 404 for a
            # GET continuation carrying searchID/searchResultPosition query params.
            # When attendance_only is set, major=1 asks the device for attendance
            # events only (a client-side filter is applied as a safety net below).
            body = {
                "AcsEventCond": {
                    "searchID": search_id,
                    "searchResultPosition": position,
                    "maxResults": max_results,
                    "major": 1 if attendance_only else 0,
                    "minor": 0,
                    "startTime": self._fmt_hik_time(start),
                    "endTime": self._fmt_hik_time(end),
                }
            }
            resp = self.post("/ISAPI/AccessControl/AcsEvent?format=json",
                             json.dumps(body), content_type="application/json")

            page = self._parse_acs_event_json(resp.text)
            events.extend(page["events"])
            self.skipped_events += page.get("skipped", 0)

            if total_matches is None:
                total_matches = page["total_matches"]
            num_on_page = page["num_of_matches"]
            if total_matches is None:
                total_matches = num_on_page

            position += num_on_page if num_on_page > 0 else len(page["events"])
            if position <= 0:
                break
            if total_matches is not None and position >= total_matches:
                break
            if num_on_page == 0 and not page["events"]:
                break
            if len(events) >= total_matches:
                break
            # Safety cap: avoid an infinite loop on misbehaving devices.
            if position >= max_results * 20:
                self.truncated_cap_hit = True
                print(f"WARNING: pagination safety cap reached after {len(events)} events; "
                      f"results may be truncated (device may have more).", file=sys.stderr)
                break

        if attendance_only:
            kept = [e for e in events if e["major"] == 1]
            self.filtered_events += len(events) - len(kept)
            return kept
        return events

    @staticmethod
    def _parse_acs_event_json(json_text):
        """Mirrors ParseAcsEventJson: AcsEvent.InfoList / AcsEvent.EventList / InfoList / EventList."""
        events = []
        try:
            doc = json.loads(json_text)
        except Exception:
            return {"events": events, "num_of_matches": 0, "total_matches": None, "skipped": 0}

        if not isinstance(doc, dict):
            return {"events": events, "num_of_matches": 0, "total_matches": None, "skipped": 0}

        root = doc.get("AcsEvent") if isinstance(doc.get("AcsEvent"), dict) else doc

        event_list = None
        for key in ("InfoList", "EventList"):
            if isinstance(root.get(key), list):
                event_list = root[key]
                break
        if event_list is None:
            for key in ("InfoList", "EventList"):
                if isinstance(doc.get(key), list):
                    event_list = doc[key]
                    break

        def as_int(v):
            try:
                return int(str(v).strip())
            except (ValueError, TypeError):
                return 0

        num_of_matches = as_int(root.get("numOfMatches")) if root else 0
        total = root.get("totalMatches") if root else None
        total_matches = as_int(total) if total is not None else None

        skipped = 0
        if isinstance(event_list, list):
            for item in event_list:
                if not isinstance(item, dict):
                    skipped += 1
                    continue
                employee_code = item.get("employeeNo") or item.get("EmployeeNo") or ""
                time_str = item.get("time") or item.get("eventTime") or ""
                try:
                    event_time = _parse_hik_time(time_str)
                except Exception:
                    skipped += 1
                    continue
                major = as_int(item.get("major"))
                minor = as_int(item.get("minor"))
                event_type = _classify_event(major, minor)
                verify_mode = item.get("verifyMode") or item.get("VerifyMode")
                in_out_mode = item.get("inAndOutMode") or item.get("InOutMode")
                events.append({
                    "employee_no": str(employee_code),
                    "time": event_time,
                    "event_type": event_type,
                    "major": major,
                    "minor": minor,
                    "verify_mode": verify_mode,
                    "in_out_mode": in_out_mode,
                    "raw": str(item)[:100] + ("..." if len(str(item)) > 100 else ""),
                })

        return {"events": events, "num_of_matches": num_of_matches, "total_matches": total_matches,
                "skipped": skipped}

    def _audit_log_search(self, since, attendance_only=False):
        """GET /ISAPI/AccessControl/AuditLog/search, mirroring ParseAuditLogXml.

        Note: this fallback targets older Hikvision firmware families that accept
        'Z'-suffixed UTC times. Newer DS-K1T3xx units reject them (404 notSupport)
        but work with the JSON AcsEvent search above.
        """
        path = "/ISAPI/AccessControl/AuditLog/search"
        if since is not None:
            s = since.strftime("%Y-%m-%dT%H:%M:%SZ")
            e = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            path += f"?searchID=1&startTime={requests.utils.quote(s)}&endTime={requests.utils.quote(e)}"
        resp = self.get(path)
        events = HikvisionClient._parse_audit_log_xml(resp.text)
        # Count raw LogItems that were dropped for consistency with the JSON path.
        try:
            root = ET.fromstring(resp.text)
            self.skipped_events += len(_find_all_local(root, "LogItem")) - len(events)
        except (ET.ParseError, ValueError, TypeError):
            pass  # Already handled (and reported as empty) by _parse_audit_log_xml.
        if attendance_only:
            kept = [e for e in events if e["major"] == 1]
            self.filtered_events += len(events) - len(kept)
            return kept
        return events

    @staticmethod
    def _parse_audit_log_xml(xml_text):
        """Mirrors ParseAuditLogXml: LogItem elements with major/minor codes."""
        events = []
        try:
            root = ET.fromstring(xml_text)
        except Exception:
            return events
        for item in _find_all_local(root, "LogItem"):
            def child(name):
                el = _find_local(item, name)
                return el.text.strip() if el is not None and el.text else ""
            employee_code = child("employeeNo")
            time_str = child("time")
            major = _as_int(child("major"))
            minor = _as_int(child("minor"))
            try:
                event_time = _parse_hik_time(time_str)
            except Exception:
                continue
            event_type = _classify_event(major, minor)
            events.append({
                "employee_no": str(employee_code),
                "time": event_time,
                "event_type": event_type,
                "major": major,
                "minor": minor,
                "verify_mode": child("VerifyMode"),
                "in_out_mode": child("inAndOutMode"),
                "raw": str(item)[:100] + ("..." if len(str(item)) > 100 else ""),
            })
        return events


def _as_int(value):
    try:
        if isinstance(value, bool):
            return int(value)
        return int(str(value).strip())
    except (ValueError, TypeError):
        return 0


def _classify_event(major, minor):
    """Map Hikvision event codes to a CheckIn/CheckOut/Other label.

    Only genuine attendance punches are classified as CheckIn/CheckOut:
      major=1, minor=75  -> CheckIn
      major=1, minor=76  -> CheckOut
    Everything else (door events, card reads, alarms, etc.) is reported as
    "Other" so access events can't be mistaken for attendance punches.
    """
    if major == 1 and minor == 75:
        return "CheckIn"
    if major == 1 and minor == 76:
        return "CheckOut"
    return "Other"


def _parse_hik_time(value):
    """Parse ISO-8601 (with or without timezone). Returns a datetime in UTC."""
    v = str(value).strip()
    if not v:
        raise ValueError("empty time")
    try:
        parsed = dt.datetime.fromisoformat(v.replace("Z", "+00:00"))
    except ValueError:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y%m%dT%H%M%S", "%Y%m%dT%H%M%S%z",
                    "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
            try:
                parsed = dt.datetime.strptime(v, fmt)
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"unrecognized time format: {v}")
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(dt.timezone.utc)


# ─────────────────────────────────────────────────────────────────────────────
# Output
# ─────────────────────────────────────────────────────────────────────────────

def export_to_excel(events, output_path):
    """Write events to an .xlsx workbook with openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Events"

    headers = ["#", "Employee No", "Time (UTC)", "Event Type", "Major", "Minor",
               "Verify Mode", "In/Out Mode", "Raw"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, evt in enumerate(events, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=evt["employee_no"])
        ws.cell(row=i + 1, column=3, value=evt["time"].strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=i + 1, column=4, value=evt["event_type"])
        ws.cell(row=i + 1, column=5, value=evt["major"])
        ws.cell(row=i + 1, column=6, value=evt["minor"])
        ws.cell(row=i + 1, column=7, value=evt["verify_mode"])
        ws.cell(row=i + 1, column=8, value=evt["in_out_mode"])
        ws.cell(row=i + 1, column=9, value=evt["raw"])

    widths = [6, 16, 22, 12, 8, 8, 14, 14, 60]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Fetch attendance events from a Hikvision ISAPI terminal.")
    parser.add_argument("--host", required=True, help="Device IP address or hostname")
    parser.add_argument("--user", default="admin", help="ISAPI username (default: admin)")
    parser.add_argument("--password", default="", help="ISAPI password")
    parser.add_argument("--port", type=int, default=None,
                        help="HTTP port (default: 80, or 443 with --https)")
    parser.add_argument("--https", action="store_true", help="Use HTTPS instead of HTTP")
    parser.add_argument("--device-info", action="store_true",
                        help="Fetch and display device info (test mode), then exit")
    parser.add_argument("--days", type=int, default=7,
                        help="Number of days to look back (default: 7)")
    parser.add_argument("--start", help="Explicit start time (ISO-8601). Overrides --days.")
    parser.add_argument("--end", help="Explicit end time (ISO-8601). Defaults to now.")
    parser.add_argument("--max-results", type=int, default=1000,
                        help="Max results per request (default: 1000)")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT,
                        help=f"Per-request read timeout in seconds (default: {DEFAULT_TIMEOUT})")
    parser.add_argument("--retries", type=int, default=MAX_RETRIES,
                        help=f"Number of retries for transient failures (default: {MAX_RETRIES})")
    parser.add_argument("--tz-offset", default=None,
                        help="Device local UTC offset, e.g. +08:00 (auto-detected from /ISAPI/System/time by default)")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx path (default: hikvision_attendance_YYYYMMDD_HHMMSS.xlsx)")
    parser.add_argument("--no-export", action="store_true",
                        help="Print events to console instead of writing an .xlsx file")
    parser.add_argument("--attendance-only", action="store_true",
                        help="Fetch attendance punches only (major=1); door/card/other "
                             "events are skipped so they don't pollute attendance records")
    args = parser.parse_args()

    if args.timeout <= 0:
        parser.error("--timeout must be a positive number of seconds")
    if args.retries < 0:
        parser.error("--retries cannot be negative")
    if args.tz_offset and not re.fullmatch(r"[+-]\d{2}:\d{2}", args.tz_offset):
        parser.error("--tz-offset must look like +08:00 or -05:00")

    port = args.port if args.port is not None else (443 if args.https else 80)
    client = HikvisionClient(args.host, args.user, args.password,
                             port=port, use_https=args.https,
                             timeout=args.timeout, max_retries=args.retries,
                             tz_offset=args.tz_offset)

    # ── Device info / connectivity test ──
    print(f"Connecting to {args.host}:{port} ...")
    try:
        info = client.fetch_device_info()
    except Exception as e:
        print(f"ERROR: Could not reach device: {e}", file=sys.stderr)
        sys.exit(1)

    print("Device reachable. Device info:")
    for k, v in info.items():
        if v:
            print(f"  {k}: {v}")

    if args.device_info:
        print("\n--device-info test passed. Exiting.")
        return

    # ── Attendance event pull ──
    if args.start:
        start_dt = _parse_hik_time(args.start)
    else:
        start_dt = dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=args.days)
    if args.end:
        end_dt = _parse_hik_time(args.end)
    else:
        end_dt = dt.datetime.now(dt.timezone.utc)

    print(f"\nFetching attendance events from {start_dt.isoformat()} to {end_dt.isoformat()} ...")
    try:
        events, source = client.fetch_attendance_events(since=start_dt,
                                                        max_results=args.max_results,
                                                        attendance_only=args.attendance_only)
    except Exception as e:
        print(f"ERROR: Failed to fetch attendance events: {e}", file=sys.stderr)
        sys.exit(1)

    events = [e for e in events if start_dt <= e["time"] <= end_dt]

    print(f"Fetched {len(events)} events via {source} endpoint.")

    if args.attendance_only and client.filtered_events:
        print(f"NOTE: {client.filtered_events} non-attendance event(s) (major != 1) "
              f"filtered out of the results.", file=sys.stderr)

    if client.skipped_events:
        print(f"NOTE: {client.skipped_events} raw event(s) skipped "
              f"(missing/unparseable time, or malformed item).", file=sys.stderr)

    if not events:
        print("No attendance events found in the requested window.")
        return

    if args.no_export:
        print("\n{:<12} {:<22} {:<10} {}".format("Employee No", "Time (UTC)", "Type", "Raw"))
        print("-" * 90)
        for e in events:
            print("{:<12} {:<22} {:<10} {}".format(
                e["employee_no"], e["time"].strftime("%Y-%m-%d %H:%M:%S"),
                e["event_type"], e["raw"]))
        return

    if not OPENPYXL_OK:
        print("openpyxl is not installed. Run: pip install openpyxl", file=sys.stderr)
        print("Showing first 20 events instead:")
        for e in events[:20]:
            print(f"  {e['time'].strftime('%Y-%m-%d %H:%M:%S')}  {e['employee_no']}  {e['event_type']}")
        sys.exit(2)

    output = args.output or dt.datetime.now().strftime("hikvision_attendance_%Y%m%d_%H%M%S.xlsx")
    try:
        path = export_to_excel(events, output)
    except Exception as e:
        print(f"ERROR: Failed to write Excel file: {e}", file=sys.stderr)
        sys.exit(1)
    print(f"Exported {len(events)} events to {path}")


if __name__ == "__main__":
    main()
